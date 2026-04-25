import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =================== 配置 ===================
SWAGGER_URL = "https://esi.evetech.net/_latest/swagger.json"
OUTPUT_FILE   = "EVE_ESI_API_全量文档.xlsx"
BASE_URL      = "https://esi.evetech.net/latest"

# ---------- 中文摘要翻译字典（可自行补充）----------
SUMMARY_CN = {

}

def translate_summary(summary):
    """返回中文摘要，若无匹配则返回原始英文"""
    return SUMMARY_CN.get(summary, summary)

# =================== 辅助函数 ===================
def load_swagger(url):
    print(f"⬇ 正在从 {url} 下载最新 Swagger 规范...")
    resp = requests.get(url, headers={"User-Agent": "ESI-Excel-Generator"})
    resp.raise_for_status()
    print("✅ 下载完成。")
    return resp.json()

def resolve_ref(ref):
    if not ref:
        return ""
    return ref.split("/")[-1]

def generate_example_url(path, parameters, public_params):
    """将路径参数替换为示例值，并附加查询参数，生成可复制的 URL"""
    url = BASE_URL + path
    query = []
    for p in parameters:
        if "$ref" in p:
            ref = resolve_ref(p["$ref"])
            p = public_params.get(ref, {})
        name = p.get("name", "")
        loc  = p.get("in", "")
        schema = p.get("schema", {})
        kind = schema.get("type", "string") if schema else "string"

        if loc == "path":
            # 根据类型给示例值
            if kind in ("integer", "number"):
                url = url.replace("{" + name + "}", "12345")
            else:
                url = url.replace("{" + name + "}", "example")
        elif loc == "query":
            if kind in ("integer", "number"):
                query.append(f"{name}=1")
            else:
                query.append(f"{name}=example")
    if query:
        url += "?" + "&".join(query)
    return url

def generate_example_response(return_type, definitions):
    """基于数据模型生成一段简化的示例 JSON 字符串"""
    if not return_type:
        return ""
    model_name = return_type.rstrip("[]")
    model = definitions.get(model_name)
    if not model:
        return ""

    def build_example(model_name):
        m = definitions.get(model_name, {})
        props = m.get("properties", {})
        ex = {}
        for field, prop in props.items():
            ptype = prop.get("type", "string")
            if "$ref" in prop:
                ex[field] = build_example(resolve_ref(prop["$ref"]))
            elif ptype == "array":
                items = prop.get("items", {})
                if "$ref" in items:
                    ex[field] = [build_example(resolve_ref(items["$ref"]))]
                else:
                    ex[field] = [f"<{items.get('type', 'object')}>"]
            elif ptype == "integer":
                ex[field] = 0
            elif ptype == "number":
                ex[field] = 0.0
            elif ptype == "boolean":
                ex[field] = False
            else:
                ex[field] = f"<{ptype}>"
        return ex

    example = build_example(model_name)
    if return_type.endswith("[]"):
        example = [example]
    return json.dumps(example, indent=2, ensure_ascii=False)

def param_text(details, public_params):
    """将参数列表格式化为可读字符串"""
    params = []
    for p in details.get("parameters", []):
        if "$ref" in p:
            ref = resolve_ref(p["$ref"])
            orig = public_params.get(ref, {})
            params.append(
                f"{orig.get('name', ref)} ({orig.get('in')} | {orig.get('type','')} | {'必填' if orig.get('required') else '选填'})"
            )
        else:
            schema = p.get("schema", {})
            params.append(
                f"{p.get('name')} ({p.get('in')} | {schema.get('type','')} | {'必填' if p.get('required',False) else '选填'})"
            )
    return "; ".join(params)

def scope_text(details):
    """提取授权范围字符串"""
    scopes = []
    for sec in details.get("security", []):
        for scheme, scope_list in sec.items():
            scopes.extend(scope_list)
    return ", ".join(scopes) if scopes else "无需授权"

# =================== 主函数：生成单个工作表 ===================
def generate_excel(swagger, output_file):
    public_params = swagger.get("parameters", {})
    paths = swagger.get("paths", {})
    definitions = swagger.get("definitions", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "ESI API 全集"

    # ---------- 定义表头 ----------
    headers = [
        "模块",              # A
        "方法",              # B
        "路径",              # C
        "中文说明",          # D
        "原始摘要 (英文)",   # E
        "描述 (英文)",       # F
        "操作 ID",           # G
        "缓存时间 (秒)",     # H
        "授权范围",          # I
        "请求参数",          # J
        "示例请求 URL",      # K
        "示例响应 JSON"      # L
    ]
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # ---------- 填充数据 ----------
    for path, methods in paths.items():
        for method, details in methods.items():
            # 一个端点可能属于多个标签，我们用分号连接
            tags = details.get("tags", [])
            module_str = ", ".join(tags) if tags else "未分类"

            summary = details.get("summary", "")
            description = details.get("description", "").replace("\n", " ")

            # 中文说明优先用翻译，无翻译则用英文
            cn_explanation = translate_summary(summary) if summary else ""

            # 示例 URL 和示例响应
            example_url = generate_example_url(path, details.get("parameters", []), public_params)

            # 响应类型（取自 200）
            resp_200 = details.get("responses", {}).get("200", {})
            schema = resp_200.get("schema")
            return_type = ""
            if schema:
                if "$ref" in schema:
                    return_type = resolve_ref(schema["$ref"])
                elif schema.get("type") == "array":
                    items = schema.get("items", {})
                    if "$ref" in items:
                        return_type = resolve_ref(items["$ref"]) + "[]"
                    else:
                        return_type = f"{items.get('type', 'object')}[]"
                else:
                    return_type = schema.get("type", "object")
            example_json = generate_example_response(return_type, definitions)

            # 参数字符串
            param_str = param_text(details, public_params)

            # 授权范围
            auth_str = scope_text(details)

            row_data = [
                module_str,
                method.upper(),
                path,
                cn_explanation,
                summary,
                description[:300],
                details.get("operationId", ""),
                details.get("x-cached-seconds", ""),
                auth_str,
                param_str,
                example_url,
                example_json
            ]
            ws.append(row_data)

    # ---------- 启用自动化筛选 ----------
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---------- 调整列宽和样式 ----------
    col_widths = {
        "A": 18, "B": 6, "C": 45, "D": 28, "E": 30,
        "F": 32, "G": 28, "H": 12, "I": 20, "J": 42, "K": 60, "L": 55
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 为数据区域设置自动换行和边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    # ---------- 冻结首行 ----------
    ws.freeze_panes = "A2"

    # ---------- 可选：添加数据模型工作表 ----------
    ws_def = wb.create_sheet("数据模型 (Definitions)")
    def_headers = ["模型名称", "字段名", "类型", "必填", "描述"]
    ws_def.append(def_headers)
    for col in range(1, len(def_headers) + 1):
        cell = ws_def.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 收集用到的数据模型
    used_models = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):  # 示例列
        example_cell = row[0]
        if example_cell.value:
            try:
                # 从示例 JSON 解析出模型？更稳妥的方法是从 return_type 列获取
                pass
            except:
                pass
    # 简便方法：重新遍历 paths，收集响应中引用的模型
    for path, methods in paths.items():
        for method, details in methods.items():
            resp_200 = details.get("responses", {}).get("200", {})
            schema = resp_200.get("schema")
            if schema:
                if "$ref" in schema:
                    used_models.add(resolve_ref(schema["$ref"]))
                elif schema.get("type") == "array":
                    items = schema.get("items", {})
                    if items and "$ref" in items:
                        used_models.add(resolve_ref(items["$ref"]))

    for model_name in sorted(used_models):
        model = definitions.get(model_name, {})
        props = model.get("properties", {})
        if not props:
            ws_def.append([model_name, "", "", "", model.get("description", "")])
        else:
            for field, prop in props.items():
                ptype = prop.get("type", "")
                if "$ref" in prop:
                    ptype = resolve_ref(prop["$ref"])
                elif ptype == "array":
                    items = prop.get("items", {})
                    if items and "$ref" in items:
                        ptype = resolve_ref(items["$ref"]) + "[]"
                    else:
                        ptype = f"{items.get('type', 'object')}[]"
                required = "是" if field in model.get("required", []) else "否"
                ws_def.append([model_name, field, ptype, required, prop.get("description", "")])

    # 定义样式
    for row in ws_def.iter_rows(min_row=2, max_row=ws_def.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
    ws_def.column_dimensions['A'].width = 30
    ws_def.column_dimensions['B'].width = 22
    ws_def.column_dimensions['C'].width = 20
    ws_def.column_dimensions['D'].width = 10
    ws_def.column_dimensions['E'].width = 60

    wb.save(output_file)
    print(f"🎉 文档生成完毕：{output_file}")

if __name__ == "__main__":
    swagger = load_swagger(SWAGGER_URL)
    generate_excel(swagger, OUTPUT_FILE)